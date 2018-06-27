#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Xiang Wang @ 2018-05-29 13:28:28


import ipdb
import logging
import random
import json

logging.basicConfig(level=logging.INFO,
                format='%(filename)s[line:%(lineno)d] %(levelname)s %(message)s',
                datefmt='%a, %d %b %Y %H:%M:%S',
                filename='./log.log',
                filemode='a')
import xlrd
from xlrd.xldate import xldate_as_datetime
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import re
import os
import pprint

al = Alignment(horizontal="center", vertical="center")

home_path = "结果"


fields = ['序号', '户主姓名', '无户主身份证号', '无户主电话', '家庭人数', '农户经济条件', '乡镇、街道名称', '树坪村村、居委会名称', '建成年份', '详细地址', '建筑面积', '建筑占地面积', '建筑层数', '砌筑方式', '结构类型', '房屋周边情况', '房屋使用情况', '私自加层', '地下室', ' （三无）无圈梁、无　地梁、无构造柱', ' 预制多孔板', ' 用地手续', '建设规划手续', '不动产权证或房屋产权证', '享受农村困难家庭危房改造政策', '享受农村困难家庭危房改造政策年份', '中央补助资金', '省级补助资金', '地方补助资金', '鉴定等级', '治理改造方式', '备注', '登记人员名称', '登记时间', 'GIS-X（经度）', 'GIS-Y（纬度）']

class HouseManager(object):
    def __init__(self, entry, country, number, manager_name):
        self.entry = entry
        self.country = country
        self.number = number
        self.manager_name = str(manager_name)
        self._data = None

    def __str__(self):
        return "HouseManager: {}-{}".format(self.country.name, self.manager_name)

    @property
    def data(self):
        if self._data is not None:
            return self._data
        else:
            self._data = self._get_data()
            return self._data

    def _get_data(self):
        wb = load_workbook(self.entry)
        ws = wb.active
        data = {
            "序号": self.number,
            "户主姓名": ws["B4"].value,
            "无户主身份证号": ws["B5"].value,
            "无户主电话": ws["I4"].value,
            "身份证号": ws["A5"].value,
            "家庭人数": ws["I5"].value,
            "乡镇、街道名称": "{}{}".format(
                self.country.town.name, self.country.name),
            "树坪村村、居委会名称": ws["B14"].value,
            "详细地址": "温州市瑞安市{}{}".format(
                self.country.town.name, self.country.name),
            "建筑面积": "{}㎡".format(ws["B22"].value),
            "建筑占地面积": "{}㎡".format(ws["I22"].value),
            "中央补助资金": ws["L44"].value,
            "省级补助资金": ws["L45"].value,
            "地方补助资金": ws["L46"].value,
            "备注": "",
            "登记人员名称": ws["B60"].value,
            "登记时间": ws["I60"].value,
        }
        assert ws['A3'].value == '户主信息表'
        assert ws['H4'].value == '电话*'
        assert ws['A5'].value == '身份证号*'
        assert ws["H5"].value == "家庭人数*"
        assert ws["A6"].value == "农户经济条件\n（单选）*"
        assert ws["B34"].value == "农家乐、民宿）"
        assert ws["B36"].value == "文化礼堂、幼儿园）"
        assert ws["A61"].value == "审核人员"

        for i in range(6, 12):
            if ws["B{}".format(i)].value == 'R':
                data["农户经济条件"] = ws["C{}".format(i)].value
                break
        else:
            logging.error("{}没有农村经济条件".format(self))

        for i in range(15, 22):
            if ws["B{}".format(i)].value == 'R':
                data["建成年份"] = ws["C{}".format(i)].value
                break
        else:
            logging.error("{}没有建成年份".format(self))

        for i in range(16, 21):
            if ws["I{}".format(i)].value == "R":
                data["建筑层数"] = ws["J{}".format(i)].value
                break
        else:
            logging.error("{}没有建筑层数".format(self))

        if ws["B23"].value == "R":
            data["砌筑方式"] = ws["C23"].value
        elif ws["D23"].value == "R":
            data["砌筑方式"] = ws["E23"].value
        else:
            logging.error("{}没有砌筑方式".format(self))

        for i in range(23,31):
            if ws["I{}".format(i)].value == "R":
                data["结构类型"] = ws["J{}".format(i)].value
                break
        else:
            logging.error("{}没有结构类型".format(self))

        for i in range(31,38):
            if ws["I{}".format(i)].value == "R":
                data["房屋周边情况"] = ws["J{}".format(i)].value
                break
        else:
            logging.error("{}没有房屋周边情况".format(self))

        for i in range(31,38):
            if ws["I{}".format(i)].value == "R":
                data["房屋周边情况"] = ws["J{}".format(i)].value
                break
        else:
            logging.error("{}没有房屋周边情况".format(self))

        for i in range(31, 38):
            if ws["B{}".format(i)].value == "R":
                if i in [33, 35]:
                    data["房屋使用情况"] = ws["C{}".format(i)].value + ws["B{}".format(i+1)].value
                else:
                    data["房屋使用情况"] = ws["C{}".format(i)].value
                break
        else:
            logging.error("{}没有房屋使用情况".format(self))

        for i in range(38, 42):
            if ws["B{}".format(i)].value == "R":
                data["私自加层"] = ws["C{}".format(i)].value
                break
        else:
            logging.error("{}没有私自加成".format(self))

        if ws["I38"].value == "R":
            data["地下室"] = ws["J38"].value
        elif ws["K38"].value == "R":
            data["地下室"] = ws["L38"].value
        else:
            logging.error("{}没有地下室情况".format(self))

        if ws["B42"].value == "R":
            data[" （三无）无圈梁、无　地梁、无构造柱"] = ws["C42"].value
        elif ws["D42"].value == "R":
            data[" （三无）无圈梁、无　地梁、无构造柱"] = ws["E42"].value
        else:
            logging.error("{}没有 （三无）无圈梁、无　地梁、无构造柱".format(self))

        if ws["I42"].value == "R":
            data[" 预制多孔板"] = ws["J42"].value
        elif ws["K42"].value == "R":
            data[" 预制多孔板"] = ws["L42"].value
        else:
            logging.error("{}没有 预制多孔板".format(self))

        if ws["B53"].value == "R":
            data[" 用地手续"] = ws["C53"].value
        elif ws["D53"].value == "R":
            data[" 用地手续"] = ws["E53"].value
        else:
            logging.error("{}没有 用地手续".format(self))

        if ws["I53"].value == "R":
            data["建设规划手续"] = ws["J53"].value
        elif ws["K53"].value == "R":
            data["建设规划手续"] = ws["L53"].value
        else:
            logging.error("{}没有建设规划手续".format(self))

        if ws["B54"].value == "R":
            data["不动产权证或房屋产权证"] = ws["C54"].value
        elif ws["D54"].value == "R":
            data["不动产权证或房屋产权证"] = ws["E54"].value
        else:
            logging.error("{}没有不动产权证或房屋产权证".format(self))

        if ws["B43"].value == "R":
            data["享受农村困难家庭危房改造政策"] = ws["C43"].value
            data["享受农村困难家庭危房改造政策年份"] = "无"
        elif ws["D43"].value == "R":
            data["享受农村困难家庭危房改造政策"] = ws["E43"].value
            for i in range(43,48):
                if ws["E{}".format(i)].value == "R":
                    data["享受农村困难家庭危房改造政策年份"] = ws["F{}".format(i)].value + "年"
                    break
            else:
                logging.error("{}没有享受农村困难家庭危房改造政策年份".format(self))
        else:
            logging.error("{}没有享受农村困难家庭危房改造政策".format(self))

        for i in range(48, 53):
            if ws["B{}".format(i)].value == "R":
                data["鉴定等级"] = ws["C{}".format(i)].value
                break
        else:
            logging.error("{}没有鉴定等级".format(self))

        for i in range(48, 53):
            if ws["I{}".format(i)].value == "R":
                data["治理改造方式"] = ws["J{}".format(i)].value
                break
        else:
            data["治理改造方式"] = ""

        location = ws["B59"].value
        longitude, latitude = re.match(
            r'^GIS-X: ([0-9.]*)       GIS-Y: ([0-9.]*)$',
            location).groups()
        data["GIS-X（经度）"] = longitude
        data["GIS-Y（纬度）"] = latitude
        return data


class Villiage(object):
    def __init__(self, entry, town):
        self.entry = entry
        self.name = entry.name
        self.town = town
        self.managers = []
        self.load_managers()

    def load_managers(self):
        # logging.info("{}载入用户".format(self))
        for manager_entry in os.scandir(self.entry):
            number, manager_name = re.match(
                r'(\d+)\s*(.*)\.xlsx$', manager_entry.name).groups()
            housemanager = HouseManager(
                entry=manager_entry,
                country=self,
                number=int(number),
                manager_name=manager_name)
            self.managers.append(housemanager)
        self.managers = sorted(self.managers, key=lambda x: x.number)

    def create_wb(self):
        logging.info("{}创建wb".format(self))
        wb = Workbook()
        ws = wb.active
        for indexi, field in enumerate(fields):
            continue
            logging.info("处理列{}".format(field))
            ws.cell(row=1, column=indexi+1, value=field)
            for indexj, manager in enumerate(self.managers):
                logging.info("写入{}".format(manager))
                ws.cell(row=indexj+2, column=indexi+1, value=manager.data[field])
        for index, field in enumerate(fields):
            ws.cell(row=1, column=index+1, value=field)
        for indexj, manager in enumerate(self.managers):
            logging.info("写入{}".format(manager))
            for indexi, field in enumerate(fields):
                logging.debug("写入field{}".format(field))
                ws.cell(row=indexj+2, column=indexi+1, value=manager.data[field])
        return wb

    def get_destiny(self):
        if os.path.isdir(os.path.join(home_path, self.town.name)) is False:
            os.mkdir(os.path.join(home_path, self.town.name))
        return os.path.join(home_path, self.town.name, self.entry.name+".xlsx")

    def __str__(self):
        return "Villiage({}-{})".format(self.town.name, self.entry.name)

    def run(self):
        logging.info("当前运行{}".format(self.entry))
        wb = self.create_wb()
        wb.save(self.get_destiny())


class Town(object):
    def __init__(self, entry):
        self.entry = entry
        self.name = entry.name
        self.villiages = self.get_villiages()

    def get_villiages(self):
        villiages = []
        for country_entry in os.scandir(self.entry):
            villiage = Villiage(
                country_entry,
                town=self)
            villiages.append(villiage)
        return villiages

    def run(self):
        logging.info("当前运行镇: {}".format(self))
        for villiage in self.villiages:
            villiage.run()

    def __str__(self):
        return "Town: {}".format(self.entry.name)


class Work(object):

    def __init__(self, path):
        logging.info("开始运行")
        self.path = path
        self.towns = []

    def run(self, town=None):
        for entry in filter(
                lambda x: x.is_dir(), os.scandir(self.path)):
            if not re.match(r'.*镇$', entry.name):
                continue
            logging.info("正在预处理{}".format(entry.name))
            town_obj = Town(entry)
            town_obj.run()


def test():
    town = Town(next(filter(lambda x: '镇' in x.name, os.scandir('./'))))
    villiage = random.choice(town.villiages)
    manager = random.choice(villiage.managers)
    # logging.info(json.dumps(
    #     manager.data, indent=4, ensure_ascii=False))


if __name__ == '__main__':
    Work(path="./").run()

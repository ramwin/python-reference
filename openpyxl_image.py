#!/usr/bin/env python3
# -*- coding: utf-8 -*-
# Xiang Wang @ 2018-05-29 13:28:28


from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Alignment
import re
import os

wb = Workbook()
ws = wb.active

al = Alignment(horizontal="center", vertical="center")

class Photoer(object):
    def __init__(self, filepath):
        self.path = filepath
        self.wb = load_workbook(self.path)
        print("载入文件: {}".format(self.path))
        self.ws = self.wb.active
        """
            "户主姓名": {
                "photer": "摄影人",
                "time": "2018-05-05 11:11:11"
            }
        """
        self.data = {
        }
        self.read_data()

    def read_data(self):
        print("读取文件")
        for index, row in enumerate(self.ws.rows):
            if index == 0:
                self.keys = self.get_keys(row)
                continue
            self.data[row[self.keys["户主姓名"]].value] = {
                "photer": row[self.keys["登记人员名称"]].value,
                "time": row[self.keys["登记时间"]].value,
            }
        print(self.data)

    def get_keys(self, row):
        """
        返回 {"户主姓名": 1, "登记人员名称": 2}...
        """
        print("get_keys")
        keys = {}
        for index, cell in enumerate(row):
            print(cell.value)
            keys[cell.value] = index
        print(keys)
        return keys


class HouseManager(object):
    def __init__(self, country_name, number, manager_name):
        self.country_name = country_name
        self.number = number
        self.manager_name = manager_name
        self.path = os.path.join(self.country_name, self.number+self.manager_name)
        self.number_int = int(number)

    def get_images(self):
        """返回图片路径列表"""
        return map(lambda x: os.path.join(self.path, x), os.listdir(self.path))


dir_name = list(filter(lambda x: os.path.isdir(x), os.listdir()))[0]
managers = []
for i in os.listdir(dir_name):
    number, manager_name = re.match(r'(\d+)(\S+)', i).groups()
    housemanager = HouseManager(
        country_name = "半岭堂村", number=number, manager_name=manager_name)
    managers.append(housemanager)
managers = sorted(managers, key=lambda x: x.number_int)


excel_file = list(filter(
    lambda x: os.path.isfile(x) and x.endswith('xlsx'),
    os.listdir()
))[0]

photoer = Photoer(excel_file)


row = 2
for index, manager in enumerate(managers):
    ws.merge_cells(
        start_row=row, start_column=2, end_row=row, end_column=8)
    ws.cell(row=row, column=2).alignment = al
    ws.cell(column=2, row=row, value="户主: {}".format(manager.manager_name))
    row += 2
    for image in manager.get_images():
        _image = Image(image)
        _image.width = 500
        _image.height = 370
        try:
            ws.add_image(_image, 'B{}'.format(row))
        except Exception as e:
            print(e)
        row += 20
    ws.merge_cells(
        start_row=row, start_column=3, end_row=row, end_column=4)
    ws.merge_cells(
        start_row=row, start_column=5, end_row=row, end_column=7)
    ws.cell(column=3, row=row, value="摄影者: {}".format(photoer.data[manager.manager_name]["photer"]))
    ws.cell(column=5, row=row, value="摄影时间: {}".format(photoer.data[manager.manager_name]["time"]))
    row += 4

wb.save("../result.xlsx")
